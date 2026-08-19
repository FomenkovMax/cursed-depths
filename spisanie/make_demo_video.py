#!/usr/bin/env python3
"""Демо-видео работы скриптов списания на вымышленных данных.

Сценарий (сверху вниз = по времени):
    1. Исходный реестр СИД (реестр для работы)         ~4.0 с
    2. Как работает скрипт 1 (реальный запуск)          ~5.0 с
    3. Результат: вкладки К списанию / Отказ / Ручное   ~3×2.6 с
    4. Переход к ст. 46
    5. Исходный реестр ст. 46 (реестр для работы)       ~4.0 с
    6. Как работает скрипт 2 (реальный запуск)          ~5.0 с
    7. Результат: вкладки К списанию / Отказ / Ручное   ~3×2.6 с

Ничего не придумывает: оба скрипта реально запускаются, реестры и результаты
на экране — настоящие сгенерированные .xlsx, прочитанные через openpyxl.
Постановочный только эффект печати команды и монтажный переход между блоками.

Запуск:  python3 make_demo_video.py [итоговый.mp4]
Нужно:   pip install pillow imageio-ffmpeg openpyxl
"""

from __future__ import annotations

import datetime as dt
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).parent
FONT_DIR = Path("/usr/share/fonts/truetype/dejavu")
MONO = FONT_DIR / "DejaVuSansMono.ttf"
MONO_BOLD = FONT_DIR / "DejaVuSansMono-Bold.ttf"

W, H = 1600, 900
FPS = 30

# Палитра — та же, что в variant-a-aurora (тёмный фон + мятный акцент),
# чтобы демо-ролик выглядел частью того же комплекта, что и презентация.
BG = (5, 9, 8)
PANEL = (10, 16, 14)
BORDER_C = (32, 48, 42)
ACCENT = (53, 214, 160)      # --mint
PROMPT_C = (63, 224, 255)    # --cyan
TEXT = (219, 232, 227)
DIM = (130, 150, 143)
GOOD = (110, 214, 150)       # Списание
BAD = (232, 130, 120)        # Отказ
WARN = (230, 190, 110)       # Ручное рассмотрение

# Те же цвета, что в spisanie/spisanie/report.py — единый источник правды
# для того, «как это выглядит в реальном Excel».
XLSX_HEADER = (0x1F, 0x3B, 0x2C)
XLSX_FILL = {
    "Списание": (0xDF, 0xF3, 0xE3),
    "Отказ": (0xFD, 0xE7, 0xE7),
    "Ручное рассмотрение": (0xFF, 0xF4, 0xD6),
}
XLSX_TEXT_DARK = (20, 30, 26)
XLSX_ROW_ALT = (244, 245, 240)   # зебра для «сырого» реестра без решений

# Порядок вкладок — как реально называются листы в report.py.
TABS = ["Сводка", "Реестр решений", "К списанию", "Отказ", "Ручное рассмотрение"]
TAB_ACCENT = {"К списанию": GOOD, "Отказ": BAD, "Ручное рассмотрение": WARN}
# «Решение» на строке соответствует названию вкладки, но сформулировано иначе
# (там — SPISANIE/OTKAZ/RUCHNIK из common.py, тут — заголовок листа).
TAB_TO_RESHENIE = {"К списанию": "Списание", "Отказ": "Отказ", "Ручное рассмотрение": "Ручное рассмотрение"}

SID_COLUMNS = ["Номер ИД", "Номер ИП", "Дата окончания", "Дата ИД",
               "ФИО Клиента", "Дата рождения клиента", "Код кредита", "Код Клиента"]
ST46_COLUMNS = SID_COLUMNS + ["Дата возбуждения ИП"]
COL_WEIGHT = {
    "Номер ИД": 1.15, "Номер ИП": 1.25, "Дата окончания": 1.0, "Дата ИД": 0.95,
    "ФИО Клиента": 1.75, "Дата рождения клиента": 1.15, "Код кредита": 1.05,
    "Код Клиента": 1.0, "Дата возбуждения ИП": 1.25,
}


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(str(MONO_BOLD if bold else MONO), size)


def fmt_cell(value: object) -> str:
    """Показать значение ячейки так, как его увидел бы человек в Excel."""
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%d.%m.%Y")
    return str(value)


# ---------------------------------------------------------------------------
# 1. Реальный прогон команд (не попадает в кадр — готовит данные заранее)
# ---------------------------------------------------------------------------

@dataclass
class Command:
    argv: list[str]
    output: str = ""

    @property
    def line(self) -> str:
        # В кадре показываем «python3», а не полный путь sys.executable —
        # выполняется всё равно точным интерпретатором, просто некрасиво в видео.
        display = ["python3" if a == sys.executable else a for a in self.argv]
        return "$ " + " ".join(display)


def run(cmd: Command) -> Command:
    result = subprocess.run(cmd.argv, cwd=ROOT, capture_output=True, text=True, timeout=180)
    text = result.stdout.strip()
    if result.stderr.strip():
        text = (text + "\n" + result.stderr.strip()).strip()
    cmd.output = text
    return cmd


def prepare() -> tuple[Command, Command]:
    """Пересоздать тестовые данные с нуля и реально прогнать оба скрипта."""
    demo_dir = ROOT / "test_data"
    if demo_dir.exists():
        shutil.rmtree(demo_dir)

    gen = Command([sys.executable, "make_test_data.py", "test_data"])
    run(gen)
    print(f"[real-run] {gen.line}\n{gen.output}\n")

    sid = Command([sys.executable, "run_sid.py", "test_data/reestr_sid.xlsx",
                   "--on-date", "01.09.2026", "-o", "test_data/out_sid.xlsx"])
    run(sid)
    print(f"[real-run] {sid.line}\n{sid.output}\n")

    st46 = Command([sys.executable, "run_st46.py", "test_data/reestr_st46.xlsx",
                    "test_data/postanovleniya", "-o", "test_data/out_st46.xlsx"])
    run(st46)
    print(f"[real-run] {st46.line}\n{st46.output}\n")

    return sid, st46


# ---------------------------------------------------------------------------
# 2. Кадры
# ---------------------------------------------------------------------------

@dataclass
class Frame:
    image: Image.Image
    seconds: float


class Reel:
    def __init__(self) -> None:
        self.frames: list[Frame] = []
        self.lines: list[tuple[str, tuple[int, int, int]]] = []

        # геометрия терминальной панели
        self.pad = 64                                       # отступ панели от краёв кадра
        self.text_x = self.pad + 32
        self.title_bar_h = 46
        self.body_top = self.pad + self.title_bar_h + 20    # с запасом от заголовка окна
        self.line_h = 30
        self.font_size = 19
        self.font = font(self.font_size)
        self.font_b = font(self.font_size, bold=True)

    # --- терминал ---
    def _canvas(self) -> tuple[Image.Image, ImageDraw.ImageDraw]:
        img = Image.new("RGB", (W, H), BG)
        draw = ImageDraw.Draw(img)
        x0, y0, x1, y1 = self.pad, self.pad, W - self.pad, H - self.pad
        draw.rounded_rectangle([x0, y0, x1, y1], radius=18, fill=PANEL, outline=BORDER_C, width=2)
        draw.rounded_rectangle([x0, y0, x1, y0 + 46], radius=18, fill=(14, 22, 19))
        draw.rectangle([x0, y0 + 28, x1, y0 + 46], fill=(14, 22, 19))
        for i, c in enumerate([(232, 106, 96), (240, 189, 79), (97, 194, 112)]):
            draw.ellipse([x0 + 22 + i * 26, y0 + 16, x0 + 22 + i * 26 + 14, y0 + 30], fill=c)
        draw.text((x0 + (x1 - x0) / 2, y0 + 23), "spisanie — терминал",
                  font=self.font, fill=DIM, anchor="mm")
        return img, draw

    def _max_chars(self) -> int:
        return (W - self.pad - self.text_x) // self._char_w()

    def _char_w(self) -> int:
        bbox = self.font.getbbox("M")
        return max(1, bbox[2] - bbox[0])

    def _wrapped(self, text: str, max_chars: int) -> list[str]:
        out: list[str] = []
        for raw in text.split("\n"):
            if not raw:
                out.append("")
                continue
            while len(raw) > max_chars:
                out.append(raw[:max_chars])
                raw = "  " + raw[max_chars:]
            out.append(raw)
        return out

    def _render(self, extra: list[tuple[str, tuple[int, int, int]]]) -> Image.Image:
        img, draw = self._canvas()
        body_h = (H - self.pad) - self.body_top
        visible = max(1, body_h // self.line_h)
        shown = (self.lines + extra)[-visible:]
        y = self.body_top
        for text, color in shown:
            draw.text((self.text_x, y), text, font=self.font, fill=color)
            y += self.line_h
        return img

    def type_command(self, text: str, cps: float) -> float:
        """Печать команды посимвольно. Возвращает потраченное время (для бюджета сцены)."""
        step = 1 / cps
        for i in range(1, len(text) + 1):
            self.frames.append(Frame(self._render([(text[:i] + "▏", PROMPT_C)]), step))
        self.frames.append(Frame(self._render([(text, PROMPT_C)]), 0.35))
        self.lines.append((text, PROMPT_C))
        return len(text) * step + 0.35

    def reveal_output(self, text: str, per_line: float, hold: float) -> None:
        lines = self._wrapped(text, self._max_chars()) if text else []
        for line in lines:
            color = self._colorize(line)
            self.lines.append((line, color))
            self.frames.append(Frame(self._render([]), per_line))
        self.lines.append(("", TEXT))
        if lines:
            self.frames.append(Frame(self._render([]), hold))

    @staticmethod
    def _colorize(line: str) -> tuple[int, int, int]:
        low = line.strip().lower()
        if low.startswith("списание:"):
            return GOOD
        if low.startswith("отказ:"):
            return BAD
        if low.startswith("ручное рассмотрение:") or low.startswith("не прочитано") \
                or low.startswith("постановления без строки"):
            return WARN
        if line.startswith("  "):
            return DIM
        return TEXT

    def code_demo(self, cmd: Command, total: float = 5.0, cps: float = 42) -> None:
        """Печать команды + вывод, уложенные в заданный бюджет времени (~5 с)."""
        self.lines = []  # каждая демонстрация кода начинается с чистого терминала
        typing = self.type_command(cmd.line, cps)
        lines = self._wrapped(cmd.output, self._max_chars()) if cmd.output else []
        per_line = 0.035
        reveal = per_line * len(lines)
        hold = max(0.8, total - typing - reveal)
        self.reveal_output(cmd.output, per_line=per_line, hold=hold)

    # --- титульные и переходные карточки ---
    def title_card(self, lines: list[tuple[str, tuple[int, int, int]]], seconds: float) -> None:
        img = Image.new("RGB", (W, H), BG)
        draw = ImageDraw.Draw(img)
        draw.rounded_rectangle([0, 0, W, 6], fill=ACCENT)
        y = H / 2 - (len(lines) * 50) / 2
        for i, (text, color) in enumerate(lines):
            size = 44 if i == 0 else 22
            f = font(size, bold=(i == 0))
            draw.text((W / 2, y), text, font=f, fill=color, anchor="mm")
            y += 58 if i == 0 else 34
        self.frames.append(Frame(img, seconds))

    # --- таблицы xlsx ---
    def registry_frame(self, path: Path, columns: list[str], subtitle: str, seconds: float) -> None:
        """Исходный реестр «как есть» — без решений, зебра-строки, как обычный Excel."""
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        headers = [c.value for c in ws[1]]
        idx = [headers.index(c) for c in columns]
        rows = [[r[i] for i in idx] for r in ws.iter_rows(min_row=2, values_only=True)]

        img = Image.new("RGB", (W, H), BG)
        draw = ImageDraw.Draw(img)
        x0, y0, x1, y1 = 60, 56, W - 60, H - 56
        draw.rounded_rectangle([x0, y0, x1, y1], radius=16, fill=(250, 250, 247))

        draw.text((x0 + 26, y0 + 20), f"{path.name}", font=font(22, bold=True), fill=(20, 20, 20))
        draw.text((x0 + 26, y0 + 50), subtitle, font=font(15), fill=(110, 116, 108))

        inner_x0, inner_x1 = x0 + 20, x1 - 20
        weights = [COL_WEIGHT.get(c, 1.0) for c in columns]
        total_w = sum(weights)
        col_x = [inner_x0]
        for w in weights:
            col_x.append(col_x[-1] + (inner_x1 - inner_x0) * w / total_w)

        top = y0 + 84
        hdr_h = 46
        hdr_f = font(12, bold=True)
        cell_f = font(13)
        draw.rectangle([inner_x0, top, inner_x1, top + hdr_h], fill=XLSX_HEADER)
        for cx0, cx1, label in zip(col_x, col_x[1:], columns):
            self._cell_text(draw, label, cx0 + 10, top, cx1 - cx0 - 16, hdr_h,
                             hdr_f, (255, 255, 255), lines_max=2)

        row_h = min(40, (y1 - 30 - (top + hdr_h)) / max(1, len(rows)))
        y = top + hdr_h
        for r, row in enumerate(rows):
            fill = (255, 255, 255) if r % 2 == 0 else XLSX_ROW_ALT
            draw.rectangle([inner_x0, y, inner_x1, y + row_h], fill=fill, outline=(226, 228, 220))
            for cx0, cx1, val in zip(col_x, col_x[1:], row):
                text = fmt_cell(val) or "—"
                text = self._ellipsize(draw, text, cell_f, cx1 - cx0 - 16)
                draw.text((cx0 + 10, y + row_h / 2), text, font=cell_f,
                          fill=XLSX_TEXT_DARK if val is not None else (170, 174, 166), anchor="lm")
            y += row_h

        self.frames.append(Frame(img, seconds))

    def sheet_frame(self, path: Path, sheet: str, seconds: float) -> None:
        """Кадр листа с решениями + строка вкладок внизу, как в настоящем Excel."""
        wb = load_workbook(path)
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        want = ["Строка", "ФИО Клиента", "Решение", "Обоснование"]
        idx = [headers.index(h) for h in want]
        rows = [[r[i] for i in idx] for r in ws.iter_rows(min_row=2, values_only=True)]
        rows = rows[:9]

        img = Image.new("RGB", (W, H), BG)
        draw = ImageDraw.Draw(img)
        x0, y0, x1, y1 = 60, 56, W - 60, H - 130   # снизу место под вкладки
        draw.rounded_rectangle([x0, y0, x1, y1], radius=16, fill=(250, 250, 247))

        title_f = font(22, bold=True)
        draw.text((x0 + 26, y0 + 20), f"{path.name} — «{sheet}»", font=title_f, fill=(20, 20, 20))

        col_x = [x0 + 26, x0 + 110, x0 + 520, x0 + 720, x1 - 26]
        top = y0 + 66
        # Высота строки капается сверху: при малом числе строк (например,
        # всего 2 в «Отказе») иначе растягивается на всю панель вместе
        # с шапкой, и заголовки таблицы «плывут» вниз. Пустой остаток панели
        # ниже данных — это нормально, как в настоящем Excel.
        row_h = min(58, (y1 - 30 - top) / max(1, len(rows) + 1))

        hdr_f = font(15, bold=True)
        cell_f = font(15)
        draw.rectangle([x0 + 16, top, x1 - 16, top + row_h], fill=XLSX_HEADER)
        for cx, label in zip(col_x, want):
            draw.text((cx, top + row_h / 2), label, font=hdr_f, fill=(255, 255, 255), anchor="lm")

        y = top + row_h
        for row in rows:
            fill = XLSX_FILL.get(row[2], (255, 255, 255))
            draw.rectangle([x0 + 16, y, x1 - 16, y + row_h], fill=fill, outline=(220, 222, 216))
            for cx, cx_next, val in zip(col_x, col_x[1:] + [x1 - 26], row):
                text = self._ellipsize(draw, str(val), cell_f, cx_next - cx - 14)
                draw.text((cx, y + row_h / 2), text, font=cell_f, fill=XLSX_TEXT_DARK, anchor="lm")
            y += row_h

        self._tab_strip(draw, sheet, y_top=H - 112, y_bottom=H - 60)
        self.frames.append(Frame(img, seconds))

    def _tab_strip(self, draw: ImageDraw.ImageDraw, active: str, y_top: int, y_bottom: int) -> None:
        tab_f = font(15)
        x = 60
        for name in TABS:
            w = draw.textlength(name, font=tab_f) + 34
            is_active = name == active
            fill = (255, 255, 255) if is_active else (16, 24, 21)
            draw.rounded_rectangle([x, y_top, x + w, y_bottom], radius=9, fill=fill)
            if is_active:
                accent = TAB_ACCENT.get(name, ACCENT)
                draw.rounded_rectangle([x, y_top, x + w, y_top + 5], radius=3, fill=accent)
            color = (24, 24, 24) if is_active else DIM
            draw.text((x + w / 2, (y_top + y_bottom) / 2 + 2), name, font=tab_f,
                       fill=color, anchor="mm")
            x += w + 10

    @staticmethod
    def _cell_text(draw, text, x, y, max_w, height, f, color, lines_max=1) -> None:
        words = str(text).split()
        lines: list[str] = []
        line = ""
        for word in words:
            probe = f"{line} {word}".strip()
            if draw.textlength(probe, font=f) <= max_w or not line:
                line = probe
            else:
                lines.append(line)
                line = word
                if len(lines) == lines_max:
                    break
        if line and len(lines) < lines_max:
            lines.append(line)
        lh = f.size + 3
        y0 = y + (height - lh * len(lines)) / 2
        for line in lines:
            draw.text((x, y0), line, font=f, fill=color)
            y0 += lh

    @staticmethod
    def _ellipsize(draw, text: str, f, max_w: float) -> str:
        if draw.textlength(text, font=f) <= max_w:
            return text
        while text and draw.textlength(text + "…", font=f) > max_w:
            text = text[:-1]
        return text + "…"


# ---------------------------------------------------------------------------
# 3. Сборка сценария
# ---------------------------------------------------------------------------

def build_reel(sid: Command, st46: Command) -> Reel:
    reel = Reel()

    reel.title_card(
        [
            ("Скрипты списания задолженности", ACCENT),
            ("срок исковой давности · ст. 46 ФЗ-229", DIM),
            ("демо на вымышленных данных", DIM),
        ],
        seconds=2.4,
    )

    # --- Скрипт 1: срок исковой давности ---
    reel.registry_frame(
        ROOT / "test_data" / "reestr_sid.xlsx", SID_COLUMNS,
        "Исходный реестр — списание по истечении срока предъявления ИД",
        seconds=4.0,
    )
    reel.code_demo(sid, total=5.0)
    for tab in ("К списанию", "Отказ", "Ручное рассмотрение"):
        reel.sheet_frame(ROOT / "test_data" / "out_sid.xlsx", tab, seconds=2.6)

    reel.title_card(
        [
            ("Скрипт 2", ACCENT),
            ("списание по ст. 46 ФЗ-229", DIM),
        ],
        seconds=2.0,
    )

    # --- Скрипт 2: ст. 46 ФЗ-229 ---
    reel.registry_frame(
        ROOT / "test_data" / "reestr_st46.xlsx", ST46_COLUMNS,
        "Исходный реестр — списание по ст. 46 ФЗ-229 (+ дата возбуждения ИП)",
        seconds=4.0,
    )
    reel.code_demo(st46, total=5.0)
    for tab in ("К списанию", "Отказ", "Ручное рассмотрение"):
        reel.sheet_frame(ROOT / "test_data" / "out_st46.xlsx", tab, seconds=2.6)

    reel.title_card(
        [
            ("spisanie/", ACCENT),
            ("README.md — как запускать и что менять", DIM),
        ],
        seconds=2.4,
    )
    return reel


# ---------------------------------------------------------------------------
# 4. Кодирование в mp4
# ---------------------------------------------------------------------------

def encode(reel: Reel, output: Path) -> None:
    import imageio_ffmpeg

    work = ROOT / ".demo_frames"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir()

    list_path = work / "list.txt"
    with list_path.open("w", encoding="utf-8") as listing:
        for i, frame in enumerate(reel.frames):
            path = work / f"f{i:05d}.png"
            frame.image.save(path)
            listing.write(f"file '{path.name}'\nduration {frame.seconds:.3f}\n")
        listing.write(f"file 'f{len(reel.frames) - 1:05d}.png'\n")

    ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    subprocess.run(
        [
            ffmpeg, "-y", "-f", "concat", "-safe", "0", "-i", str(list_path),
            "-vsync", "cfr", "-r", str(FPS),
            "-vf", "format=yuv420p",
            "-c:v", "libx264", "-preset", "medium", "-crf", "20",
            "-movflags", "+faststart",
            str(output),
        ],
        check=True, capture_output=True, text=True,
    )
    shutil.rmtree(work)


def main() -> int:
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "demo.mp4"

    print("== Реальный прогон команд ==")
    sid, st46 = prepare()

    print("== Рендер кадров ==")
    reel = build_reel(sid, st46)
    print(f"кадров: {len(reel.frames)}, "
          f"длительность: {sum(f.seconds for f in reel.frames):.1f} с")

    print("== Кодирование в mp4 ==")
    encode(reel, output)

    size_mb = output.stat().st_size / 1e6
    print(f"Готово: {output} ({size_mb:.1f} MB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

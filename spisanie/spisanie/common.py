"""Общие части обоих скриптов списания: чтение реестра, нормализация, даты.

Здесь сознательно нет никакой бизнес-логики про сроки — только «привести грязные
данные к пригодному виду и честно сказать, что не получилось». Решение принимают
модули sid.py (срок исковой давности) и st46.py (ст. 46 ФЗ-229).
"""

from __future__ import annotations

import datetime as dt
import re
import unicodedata
from dataclasses import dataclass, field

from dateutil.relativedelta import relativedelta

# ---------------------------------------------------------------------------
# Колонки реестра
# ---------------------------------------------------------------------------
# Ключ — внутреннее имя поля, значение — варианты заголовка в присылаемом файле.
# Реальные выгрузки называют колонки по-разному, поэтому сопоставляем по списку
# синонимов, а не по точному совпадению.
COLUMN_SYNONYMS: dict[str, list[str]] = {
    "id_nomer": ["номер ид", "no ид", "ид номер", "номер исполнительного документа"],
    "ip_nomer": ["номер ип", "no ип", "ип номер", "номер исполнительного производства"],
    "ip_data_okonchaniya": ["дата окончания", "дата окончания ип", "дата окончания исполнительного производства"],
    "id_data": ["дата ид", "дата исполнительного документа"],
    "fio": ["фио клиента", "фио", "фио должника", "клиент"],
    "data_rozhdeniya": ["дата рождения клиента", "дата рождения", "др клиента", "др"],
    "kod_kredita": ["код кредита", "номер кредита", "код договора"],
    "kod_klienta": ["код клиента", "id клиента"],
    # только для ст. 46
    "ip_data_vozbuzhdeniya": ["дата возбуждения ип", "дата возбуждения", "дата возбуждения исполнительного производства"],
}

# Человекочитаемые названия — для сообщений об ошибках и заголовков отчёта.
FIELD_TITLES: dict[str, str] = {
    "id_nomer": "Номер ИД",
    "ip_nomer": "Номер ИП",
    "ip_data_okonchaniya": "Дата окончания",
    "id_data": "Дата ИД",
    "fio": "ФИО Клиента",
    "data_rozhdeniya": "Дата рождения клиента",
    "kod_kredita": "Код кредита",
    "kod_klienta": "Код клиента",
    "ip_data_vozbuzhdeniya": "Дата возбуждения ИП",
}

DATE_FIELDS = {"ip_data_okonchaniya", "id_data", "data_rozhdeniya", "ip_data_vozbuzhdeniya"}

# Решения
SPISANIE = "Списание"
OTKAZ = "Отказ"
RUCHNIK = "Ручное рассмотрение"


class RegistryError(Exception):
    """Реестр невозможно обработать целиком (например, нет обязательной колонки)."""


# ---------------------------------------------------------------------------
# Нормализация
# ---------------------------------------------------------------------------

def norm_header(value: object) -> str:
    """Заголовок колонки -> ключ для поиска синонима."""
    text = str(value or "").replace(" ", " ")
    text = unicodedata.normalize("NFKC", text).strip().lower().replace("ё", "е")
    text = re.sub(r"[№#]", "no ", text)
    text = re.sub(r"[^\w\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_text(value: object) -> str:
    """Общая нормализация текстового значения (для сравнения, не для вывода)."""
    text = str(value or "").replace(" ", " ")
    text = unicodedata.normalize("NFKC", text).strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def norm_fio(value: object) -> str:
    """ФИО для сравнения: регистр, лишние пробелы и дефисы приводим к одному виду."""
    text = norm_text(value)
    text = text.replace("-", "-")
    return re.sub(r"\s*-\s*", "-", text)


def norm_ip_nomer(value: object) -> str:
    """Номер ИП для сравнения.

    В выгрузке и в постановлении один и тот же номер пишут по-разному:
    «12345/20/36001-ИП», «12345/20/36001 ИП», «12345/20/36001». Приводим к
    цифробуквенному ядру без разделителей и без хвоста «ип».
    """
    text = norm_text(value)
    text = re.sub(r"[\s‐-―\-–—]", "", text)
    text = re.sub(r"ип$", "", text)
    return text


def norm_kod(value: object) -> str:
    """Код кредита/клиента: убираем пробелы и хвост '.0' от чисел из Excel."""
    text = norm_text(value)
    return re.sub(r"\.0$", "", text.replace(" ", ""))


# ---------------------------------------------------------------------------
# Даты
# ---------------------------------------------------------------------------

_DATE_FORMATS = ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y")

# Excel хранит даты числом дней от 30.12.1899 (историческая ошибка с 1900 годом).
_EXCEL_EPOCH = dt.date(1899, 12, 30)


def parse_date(value: object) -> dt.date | None:
    """Разобрать дату из ячейки Excel. Возвращает None, если разобрать нельзя.

    Осознанно не угадываем: '13.13.2020' или 'нет данных' дают None, и строка
    уходит на ручное рассмотрение, а не получает случайную дату.
    """
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        # Excel-сериал. Диапазон ограничиваем, чтобы не превращать в дату
        # случайно попавший в колонку номер договора.
        number = int(value)
        if 1 <= number <= 80000:
            return _EXCEL_EPOCH + dt.timedelta(days=number)
        return None

    text = norm_text(value)
    if not text or text in {"nan", "nat", "none", "-", "нет данных", "н д"}:
        return None
    text = text.split(" ")[0]
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def years_between(start: dt.date, end: dt.date) -> float:
    """Полных лет между датами, дробью — только для вывода в отчёт."""
    return (end - start).days / 365.25


def date_plus_years(start: dt.date, years: int) -> dt.date:
    return start + relativedelta(years=years)


def date_plus_months(start: dt.date, months: int) -> dt.date:
    return start + relativedelta(months=months)


def fmt(value: dt.date | None) -> str:
    return value.strftime("%d.%m.%Y") if value else "—"


# ---------------------------------------------------------------------------
# Строка реестра
# ---------------------------------------------------------------------------

@dataclass
class Row:
    """Одна строка реестра: сырые значения, разобранные значения и проблемы."""

    number: int                      # номер строки в Excel, как её видит пользователь
    raw: dict[str, object]           # значения как есть — для вывода в отчёт
    values: dict[str, object] = field(default_factory=dict)   # разобранные
    problems: list[str] = field(default_factory=list)

    def date(self, key: str) -> dt.date | None:
        value = self.values.get(key)
        return value if isinstance(value, dt.date) else None

    def text(self, key: str) -> str:
        return str(self.raw.get(key, "") or "").strip()


@dataclass
class Decision:
    """Результат по одной строке."""

    row: Row
    decision: str            # SPISANIE | OTKAZ | RUCHNIK
    reason_code: str
    reason: str
    judgment: str = ""       # текст мотивированного суждения
    extra: dict[str, object] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Чтение реестра
# ---------------------------------------------------------------------------

def read_registry(path, required: list[str]) -> list[Row]:
    """Прочитать xlsx-реестр и разобрать строки.

    `required` — внутренние имена обязательных полей. Отсутствие колонки в файле
    считаем ошибкой всего файла (RegistryError), а пустое или неразбираемое
    значение в конкретной строке — проблемой строки.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise RegistryError("Файл реестра пуст — нет даже строки заголовков.") from exc

    # Заголовок -> внутреннее имя поля
    lookup = {syn: key for key, syns in COLUMN_SYNONYMS.items() for syn in syns}
    col_of: dict[str, int] = {}
    for index, cell in enumerate(header):
        key = lookup.get(norm_header(cell))
        if key and key not in col_of:
            col_of[key] = index

    missing = [FIELD_TITLES[k] for k in required if k not in col_of]
    if missing:
        found = ", ".join(sorted(str(c) for c in header if c)) or "—"
        raise RegistryError(
            "В реестре нет обязательных колонок: " + ", ".join(missing)
            + f".\nНайденные заголовки: {found}"
            + "\nЕсли колонка называется иначе — добавьте её название в COLUMN_SYNONYMS."
        )

    rows: list[Row] = []
    for offset, cells in enumerate(rows_iter, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in cells):
            continue  # пустые строки-разделители в конце выгрузки

        raw = {key: cells[idx] if idx < len(cells) else None for key, idx in col_of.items()}
        row = Row(number=offset, raw=raw)

        for key in required:
            value = raw.get(key)
            if value is None or str(value).strip() == "":
                row.problems.append(f"не заполнено поле «{FIELD_TITLES[key]}»")
                continue
            if key in DATE_FIELDS:
                parsed = parse_date(value)
                if parsed is None:
                    row.problems.append(
                        f"не удалось разобрать дату в поле «{FIELD_TITLES[key]}»: {value!r}"
                    )
                else:
                    row.values[key] = parsed
            else:
                row.values[key] = str(value).strip()

        rows.append(row)

    workbook.close()
    return rows


def find_duplicates(rows: list[Row]) -> dict[int, str]:
    """Найти повторы по паре (Номер ИП, Код кредита).

    Дубль в реестре — это риск списать одну и ту же задолженность дважды,
    поэтому все строки такой группы уходят на ручное рассмотрение.
    """
    groups: dict[tuple[str, str], list[int]] = {}
    for row in rows:
        key = (norm_ip_nomer(row.raw.get("ip_nomer")), norm_kod(row.raw.get("kod_kredita")))
        if not key[0] and not key[1]:
            continue
        groups.setdefault(key, []).append(row.number)

    flagged: dict[int, str] = {}
    for (ip, kredit), numbers in groups.items():
        if len(numbers) > 1:
            others = ", ".join(str(n) for n in numbers)
            for number in numbers:
                flagged[number] = (
                    f"дубль в реестре: ИП {ip or '—'} / кредит {kredit or '—'} "
                    f"встречается в строках {others}"
                )
    return flagged

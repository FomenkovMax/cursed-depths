"""Запись результата в Excel: реестр решений + разбивка по листам + сводка."""

from __future__ import annotations

import datetime as dt
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .common import FIELD_TITLES, OTKAZ, RUCHNIK, SPISANIE, Decision, fmt

HEADER_FILL = PatternFill("solid", fgColor="1F3B2C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DECISION_FILL = {
    SPISANIE: PatternFill("solid", fgColor="DFF3E3"),
    OTKAZ: PatternFill("solid", fgColor="FDE7E7"),
    RUCHNIK: PatternFill("solid", fgColor="FFF4D6"),
}
THIN = Side(style="thin", color="D4D8D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Ширины подобраны так, чтобы реестр читался без ручной подгонки;
# колонка с текстом суждения — широкая и с переносом.
WIDTHS = {"Мотивированное суждение": 96, "Обоснование": 52, "ФИО Клиента": 28}
DEFAULT_WIDTH = 18


def build_rows(decisions: list[Decision], fields: list[str]) -> tuple[list[str], list[list]]:
    """Собрать заголовки и строки: исходные поля + результат."""
    headers = (
        ["Строка"]
        + [FIELD_TITLES[f] for f in fields]
        + ["Решение", "Код причины", "Обоснование", "Мотивированное суждение"]
    )
    extra_keys: list[str] = []
    for decision in decisions:
        for key in decision.extra:
            if key not in extra_keys:
                extra_keys.append(key)
    headers += extra_keys

    body = []
    for decision in decisions:
        row = [decision.row.number]
        for f in fields:
            value = decision.row.raw.get(f)
            row.append(fmt(value) if isinstance(value, dt.date) else value)
        row += [decision.decision, decision.reason_code, decision.reason, decision.judgment]
        row += [decision.extra.get(k, "") for k in extra_keys]
        body.append(row)
    return headers, body


def _write_sheet(workbook: Workbook, title: str, headers: list[str], body: list[list]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34

    decision_col = headers.index("Решение") + 1
    for row in body:
        sheet.append(row)
        written = sheet[sheet.max_row]
        fill = DECISION_FILL.get(written[decision_col - 1].value)
        for cell in written:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if fill:
                cell.fill = fill

    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS.get(header, DEFAULT_WIDTH)
    sheet.freeze_panes = "B2"
    if body:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(body) + 1}"


def _write_summary(workbook: Workbook, decisions: list[Decision], meta: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Сводка", 0)
    sheet.column_dimensions["A"].width = 46
    sheet.column_dimensions["B"].width = 58

    def line(name: str, value: object, bold: bool = False) -> None:
        sheet.append([name, value])
        if bold:
            for cell in sheet[sheet.max_row]:
                cell.font = Font(bold=True)

    line("Отчёт", meta.get("title", ""), bold=True)
    for key, value in meta.items():
        if key != "title":
            line(key, value)
    sheet.append([])

    line("Всего строк в реестре", len(decisions), bold=True)
    for name in (SPISANIE, OTKAZ, RUCHNIK):
        count = sum(1 for d in decisions if d.decision == name)
        share = f"{count / len(decisions) * 100:.1f}%" if decisions else "—"
        line(name, f"{count}  ({share})")

    sheet.append([])
    line("Причины (код: количество)", "", bold=True)
    for code, count in Counter(d.reason_code for d in decisions).most_common():
        line(code, count)


def write_report(path, decisions: list[Decision], fields: list[str], meta: dict[str, object]) -> None:
    """Записать xlsx: Сводка, Реестр решений и три листа по видам решения."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    headers, body = build_rows(decisions, fields)
    _write_sheet(workbook, "Реестр решений", headers, body)

    decision_index = headers.index("Решение")
    for name, sheet_title in (
        (SPISANIE, "К списанию"),
        (OTKAZ, "Отказ"),
        (RUCHNIK, "Ручное рассмотрение"),
    ):
        subset = [r for r in body if r[decision_index] == name]
        _write_sheet(workbook, sheet_title, headers, subset)

    _write_summary(workbook, decisions, meta)
    workbook.save(path)

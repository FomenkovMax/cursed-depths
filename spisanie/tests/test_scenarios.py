"""Проверка обоих скриптов на вымышленном наборе данных.

Тест держит ожидаемое решение по КАЖДОЙ строке, а не только итоговые счётчики:
иначе две ошибки в разные стороны компенсируют друг друга и прогон выглядит
зелёным. Отдельно закреплены пограничные случаи (ровно 3 года, ровно 2 месяца) —
именно на них ломается формулировка «более чем».
"""

from __future__ import annotations

import datetime as dt
import subprocess
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from spisanie.common import OTKAZ, RUCHNIK, SPISANIE, parse_date, read_registry  # noqa: E402
from spisanie.postanovlenie import load_folder, parse  # noqa: E402
from spisanie.st46 import fio_match  # noqa: E402
from spisanie import sid, st46  # noqa: E402

ON_DATE = dt.date(2026, 9, 1)


@pytest.fixture(scope="session")
def data(tmp_path_factory) -> Path:
    """Сгенерировать вымышленный набор во временной папке."""
    folder = tmp_path_factory.mktemp("test_data")
    subprocess.run(
        [sys.executable, str(ROOT / "make_test_data.py"), str(folder)],
        check=True, capture_output=True, cwd=ROOT,
    )
    return folder


# ---------------------------------------------------------------------------
# Скрипт 1: срок исковой давности
# ---------------------------------------------------------------------------

# Строка Excel -> (решение, фрагмент обоснования)
SID_EXPECTED = {
    2:  (SPISANIE, "более 3 лет"),
    3:  (SPISANIE, "более 3 лет"),
    4:  (OTKAZ,    "не истёк"),
    5:  (OTKAZ,    "не истёк"),          # ровно 3 года — «более» не выполнено
    6:  (SPISANIE, "более 3 лет"),       # 3 года и 1 день
    7:  (RUCHNIK,  "не заполнено поле «Дата окончания»"),
    8:  (RUCHNIK,  "не удалось разобрать дату"),
    9:  (SPISANIE, "более 3 лет"),       # дата числом Excel
    10: (RUCHNIK,  "позже отчётной даты"),
    11: (RUCHNIK,  "дата ИД"),
    12: (RUCHNIK,  "дубль в реестре"),
    13: (RUCHNIK,  "дубль в реестре"),
    14: (RUCHNIK,  "не заполнено поле «Код клиента»"),
}


def test_sid_kazhdaya_stroka(data: Path) -> None:
    rows = read_registry(data / "reestr_sid.xlsx", sid.REQUIRED)
    decisions = {d.row.number: d for d in sid.decide_all(rows, ON_DATE)}

    assert set(decisions) == set(SID_EXPECTED), "изменился состав строк тестового реестра"
    for number, (decision, fragment) in SID_EXPECTED.items():
        got = decisions[number]
        assert got.decision == decision, f"строка {number}: {got.decision}, ждали {decision}"
        assert fragment.lower() in got.reason.lower(), f"строка {number}: {got.reason}"


def test_sid_granitsa_rovno_tri_goda(data: Path) -> None:
    """Ровно три года — это НЕ «более трёх лет»."""
    rows = read_registry(data / "reestr_sid.xlsx", sid.REQUIRED)
    decisions = {d.row.number: d for d in sid.decide_all(rows, ON_DATE)}

    rovno = decisions[5]
    assert rovno.decision == OTKAZ
    # Формулировка не должна противоречить выводу: «прошло 3 года» + отказ.
    assert "3 года" in rovno.reason
    assert "менее" not in rovno.reason

    den_spustya = decisions[6]
    assert den_spustya.decision == SPISANIE


def test_sid_vse_resheniya_imeyut_tekst_suzhdeniya(data: Path) -> None:
    rows = read_registry(data / "reestr_sid.xlsx", sid.REQUIRED)
    for decision in sid.decide_all(rows, ON_DATE):
        assert decision.judgment.strip(), f"строка {decision.row.number}: пустое суждение"
        assert "ВЫВОД:" in decision.judgment


def test_sid_otchetnaya_data_vliyaet(data: Path) -> None:
    """На более раннюю дату списаний должно быть не больше."""
    rows = read_registry(data / "reestr_sid.xlsx", sid.REQUIRED)
    pozzhe = sum(1 for d in sid.decide_all(rows, ON_DATE) if d.decision == SPISANIE)
    ranshe = sum(
        1 for d in sid.decide_all(rows, dt.date(2024, 1, 1)) if d.decision == SPISANIE
    )
    assert ranshe < pozzhe


# ---------------------------------------------------------------------------
# Скрипт 2: ст. 46 ФЗ-229
# ---------------------------------------------------------------------------

ST46_EXPECTED = {
    2:  (SPISANIE, "реквизиты сверены"),
    3:  (SPISANIE, "реквизиты сверены"),          # ФИО в родительном падеже
    4:  (OTKAZ,    "не более 2 месяцев"),
    5:  (OTKAZ,    "не более 2 месяцев"),         # ровно 2 месяца
    6:  (OTKAZ,    "дата рождения не совпадает"),
    7:  (OTKAZ,    "дата окончания ИП не совпадает"),
    8:  (OTKAZ,    "не подтверждено отсутствие имущества"),
    9:  (OTKAZ,    "п. 1 ч. 1 ст. 46"),
    10: (RUCHNIK,  "не найдено постановление"),
    11: (RUCHNIK,  "не найдено постановление"),   # скан не прочитан
    12: (SPISANIE, "реквизиты сверены"),          # .docx
}


def test_st46_kazhdaya_stroka(data: Path) -> None:
    rows = read_registry(data / "reestr_st46.xlsx", st46.REQUIRED)
    documents, _ = load_folder(data / "postanovleniya")
    decisions = {d.row.number: d for d in st46.decide_all(rows, documents)}

    assert set(decisions) == set(ST46_EXPECTED), "изменился состав строк тестового реестра"
    for number, (decision, fragment) in ST46_EXPECTED.items():
        got = decisions[number]
        assert got.decision == decision, f"строка {number}: {got.decision}, ждали {decision}"
        assert fragment.lower() in got.reason.lower(), f"строка {number}: {got.reason}"


def test_st46_skan_ne_chitaetsya_i_ne_spisyvaetsya(data: Path) -> None:
    """Скан без текстового слоя не должен молча превращаться в решение."""
    _, broken = load_folder(data / "postanovleniya")
    assert any("текстового слоя" in why for _, why in broken)


def test_st46_postanovlenie_bez_stroki_reestra(data: Path) -> None:
    rows = read_registry(data / "reestr_st46.xlsx", st46.REQUIRED)
    documents, _ = load_folder(data / "postanovleniya")
    orphans = st46.unmatched_documents(rows, documents)
    assert len(orphans) == 1
    assert "22999" in (orphans[0].ip_nomer or "")


def test_st46_granitsa_rovno_dva_mesyatsa(data: Path) -> None:
    rows = read_registry(data / "reestr_st46.xlsx", st46.REQUIRED)
    documents, _ = load_folder(data / "postanovleniya")
    decisions = {d.row.number: d for d in st46.decide_all(rows, documents)}
    assert decisions[5].decision == OTKAZ
    assert "2 месяца" in decisions[5].extra["Длительность ИП"]


def test_st46_parsing_rekvizitov(data: Path) -> None:
    """Из постановления вытащены все реквизиты, ничего не потеряно."""
    documents, _ = load_folder(data / "postanovleniya")
    document = documents["22101/24/36001"]
    assert document.data_vozbuzhdeniya == dt.date(2024, 4, 2)
    assert document.data_okonchaniya == dt.date(2024, 11, 18)
    assert document.data_rozhdeniya == dt.date(1978, 7, 14)
    assert document.st46 is True
    assert document.punkt == 4
    assert document.net_imushchestva is True
    assert document.missing == []


# ---------------------------------------------------------------------------
# Сверка ФИО
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "reestr, dokument, sovpalo",
    [
        ("Ковалёв Пётр Сергеевич", "Ковалёв Пётр Сергеевич", True),
        ("Ковалёв Пётр Сергеевич", "Ковалева Петра Сергеевича", True),   # род. падеж
        ("Астахова Мария Игоревна", "Астаховой Марии Игоревны", True),
        ("Ковалёв Пётр Сергеевич", "Ковалев Петр Сергеевич", True),      # е/ё
        ("Мещеряков Илья Борисович", "Мещерякова Ильи Борисовича", True),
        ("Иванов Иван Иванович", "Иваненко Иван Иванович", False),       # разные люди
        ("Иванов Иван Иванович", "Петров Иван Иванович", False),
        ("Иванов Иван Иванович", "Иванов Иван", False),                  # нет отчества
        ("Иванов Иван Иванович", "Иванов Игорь Иванович", False),        # другое имя
        ("Иванов Иван Иванович", "", False),
    ],
)
def test_fio_match(reestr: str, dokument: str, sovpalo: bool) -> None:
    assert fio_match(reestr, dokument)[0] is sovpalo


def test_fio_match_izvestnoe_ogranichenie() -> None:
    """Задокументированное ограничение сверки ФИО.

    «Иванова» — это одновременно родительный падеж от «Иванов» и именительный
    падеж женской фамилии. Отличить одно от другого без морфологического
    анализа нельзя, поэтому сверка считает их совпадением. Ошибку перехватывает
    независимая сверка даты рождения, а сам факт попадает в отчёт как
    примечание «сверено с учётом склонения».
    """
    sovpalo, sklonenie = fio_match("Иванов Иван Иванович", "Иванова Ивана Ивановича")
    assert sovpalo is True
    assert sklonenie is True, "случай должен быть помечен в отчёте, а не пройти молча"


# ---------------------------------------------------------------------------
# Разбор дат
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "znachenie, ozhidanie",
    [
        ("15.03.2021", dt.date(2021, 3, 15)),
        ("2021-03-15", dt.date(2021, 3, 15)),
        ("15/03/2021", dt.date(2021, 3, 15)),
        (dt.datetime(2021, 3, 15, 10, 30), dt.date(2021, 3, 15)),
        (44835, dt.date(2022, 10, 1)),          # серийный номер Excel
        ("нет данных", None),
        ("31.02.2021", None),                    # несуществующая дата
        ("", None),
        (None, None),
        (4410023, None),                         # номер договора — не дата
    ],
)
def test_parse_date(znachenie, ozhidanie) -> None:
    assert parse_date(znachenie) == ozhidanie


# ---------------------------------------------------------------------------
# Запуск через CLI — проверяем, что отчёт реально формируется
# ---------------------------------------------------------------------------

def test_cli_sid(data: Path, tmp_path: Path) -> None:
    output = tmp_path / "sid.xlsx"
    result = subprocess.run(
        [sys.executable, str(ROOT / "run_sid.py"), str(data / "reestr_sid.xlsx"),
         "--on-date", "01.09.2026", "-o", str(output)],
        capture_output=True, text=True, cwd=ROOT,
    )
    assert result.returncode == 0, result.stderr
    assert output.exists()

    from openpyxl import load_workbook

    workbook = load_workbook(output)
    assert {"Сводка", "Реестр решений", "К списанию", "Отказ", "Ручное рассмотрение"} <= set(
        workbook.sheetnames
    )
    assert workbook["Реестр решений"].max_row == len(SID_EXPECTED) + 1


def test_cli_st46(data: Path, tmp_path: Path) -> None:
    output = tmp_path / "st46.xlsx"
    result = subprocess.run(
        [sys.executable, str(ROOT / "run_st46.py"), str(data / "reestr_st46.xlsx"),
         str(data / "postanovleniya"), "-o", str(output)],
        capture_output=True, text=True, cwd=ROOT,
    )
    assert result.returncode == 0, result.stderr
    assert output.exists()
    assert "скан" in result.stdout.lower() or "текстового слоя" in result.stdout.lower()


def test_cli_ponyatnaya_oshibka_pri_nehvatke_kolonki(tmp_path: Path) -> None:
    """Если в реестре нет нужной колонки — внятное сообщение, а не трассировка."""
    from openpyxl import Workbook

    path = tmp_path / "krivoy.xlsx"
    workbook = Workbook()
    workbook.active.append(["Что-то", "Совсем другое"])
    workbook.active.append([1, 2])
    workbook.save(path)

    result = subprocess.run(
        [sys.executable, str(ROOT / "run_sid.py"), str(path)],
        capture_output=True, text=True, cwd=ROOT,
    )
    assert result.returncode == 2
    assert "нет обязательных колонок" in result.stderr
    assert "Traceback" not in result.stderr
